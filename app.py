import pandas as pd
import re
import phonenumbers
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os
from twilio.rest import Client

# === CONFIGURAZIONI ===
input_file = "TEST_INSCRITOS.xlsx"  # Cambia il nome se serve
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_folder = f"output_{timestamp}"
os.makedirs(output_folder, exist_ok=True)

# Twilio credentials (modifica con le tue)
TWILIO_ACCOUNT_SID = 'x'
TWILIO_AUTH_TOKEN = 'x'
TWILIO_PHONE_NUMBER = 'x'  # Il tuo numero Twilio
client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

# === MAPPA PAESE-CODICE ===
country_code_map = {
    "ITALIA": "IT",
    "FRANCIA": "FR",
    "BELGICA": "BE",
    "REINO UNIDO DE GRAN BRETAÑA E IRLANDA DEL NORTE": "GB",
    "ALEMANIA": "DE",
    "COREA (SUR) REPUBLICA DE": "KR",
    "SUIZA": "CH",
    "CHINA": "CN",
    "FEDERACION DE RUSIA": "RU",
    "SUECIA": "SE",
    "TURQUIA": "TR",
    "AUSTRIA": "AT",
    "ESPAÑA": "ES",
    "PAISES BAJOS (HOLANDA)": "NL",
    "HUNGRIA": "HU",
    "AUSTRALIA": "AU",
    "JAPON": "JP",
    "INDONESIA": "ID"
}

# === FUNZIONI ===

def clean_phone_number(number):
    if pd.isna(number):
        return ""
    cleaned = re.sub(r"[^\d+]", "", str(number))
    if not cleaned.startswith('+'):
        return cleaned
    for code in country_code_map.values():
        if cleaned.startswith('+' + code):
            cleaned = re.sub(r'^\+(' + code + r')00', r'+\1', cleaned)
            cleaned = cleaned[:len(code) + 1] + cleaned[len(code) + 1:].lstrip('0')
    return cleaned

def is_valid_phone_number(number, country_name):
    country_code = country_code_map.get(country_name, 'IT')
    try:
        phone_number = phonenumbers.parse(number, country_code)
        return phonenumbers.is_valid_number(phone_number)
    except phonenumbers.phonenumberutil.NumberParseException:
        return False

def color_invalid_cells(file_path, df, validation_column, no_prefix_column):
    wb = load_workbook(file_path)
    ws = wb.active
    for row in range(2, len(df) + 2):
        val_cell = ws.cell(row=row, column=df.columns.get_loc(validation_column) + 1)
        prefix_cell = ws.cell(row=row, column=df.columns.get_loc(no_prefix_column) + 1)
        if val_cell.value == 'Invalid':
            val_cell.font = Font(color="FF0000")
        if prefix_cell.value == 'No prefix':
            prefix_cell.font = Font(color="FF0000")
    wb.save(file_path)

# === ELABORAZIONE ===

print("📥 Caricamento del file...")
df = pd.read_excel(input_file)

for col in df.columns:
    if df[col].astype(str).str.contains(r'\+\(\d{1,3}\)\s?\d+', regex=True, na=False).any():
        print(f"📞 Colonna con numeri trovata: {col}")
        
        df[col] = df[col].apply(clean_phone_number)

        validation_column = col + " Valid"
        no_prefix_column = col + " No Prefix"

        df[validation_column] = df.apply(
            lambda row: 'Valid' if is_valid_phone_number(row[col], row['Paese']) and row[col].startswith('+') else 'Invalid',
            axis=1
        )
        df[no_prefix_column] = df.apply(
            lambda row: 'No prefix' if not row[col].startswith('+') else '',
            axis=1
        )
        break

valid_count = df[df[validation_column] == 'Valid'].shape[0]
invalid_count = df[df[validation_column] == 'Invalid'].shape[0]
print(f"✅ Numeri validi: {valid_count}")
print(f"❌ Numeri non validi: {invalid_count}")

# === INVIO SMS ===

sent_sms_file = os.path.join(output_folder, "sent_sms.xlsx")
if os.path.exists(sent_sms_file):
    sent_df = pd.read_excel(sent_sms_file)
    already_sent = set(sent_df['Phone'].astype(str))
else:
    sent_df = pd.DataFrame(columns=['Phone', 'Status', 'Timestamp'])
    already_sent = set()

sent_rows = []

print("📤 Invio SMS...")

for index, row in df[df[validation_column] == 'Valid'].iterrows():
    phone = row[col]
    if phone in already_sent:
        continue
    try:
        message = client.messages.create(
            body="Ciao è un test",
            from_=TWILIO_PHONE_NUMBER,
            to=phone
        )
        print(f"✅ SMS inviato a: {phone}")
        sent_rows.append({'Phone': phone, 'Status': 'Sent', 'Timestamp': datetime.now()})
    except Exception as e:
        print(f"❌ Errore con {phone}: {e}")
        sent_rows.append({'Phone': phone, 'Status': f'Error: {e}', 'Timestamp': datetime.now()})

# Salva log SMS
new_sent_df = pd.DataFrame(sent_rows)
sent_df = pd.concat([sent_df, new_sent_df], ignore_index=True)
sent_df.drop_duplicates(subset=['Phone'], inplace=True)
sent_df.to_excel(sent_sms_file, index=False)
print(f"📝 Log SMS salvato in: {sent_sms_file}")

# === SALVATAGGIO FINALE ===
output_path = os.path.join(output_folder, f"{timestamp}_cleaned.xlsx")
df.to_excel(output_path, index=False)
color_invalid_cells(output_path, df, validation_column, no_prefix_column)
print(f"✅ File finale salvato: {output_path}")
